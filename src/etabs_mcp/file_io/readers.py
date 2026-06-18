"""File readers: CSV and XLSX."""

from __future__ import annotations

import abc
import csv
from datetime import date, datetime
from datetime import time as dt_time
from pathlib import Path
from typing import Any

import chardet
import openpyxl

from etabs_mcp.file_io.const import (
    MAX_FILE_SIZE_BYTES,
    MAX_INPUT_COLUMNS,
    MAX_INPUT_ROWS,
    MAX_INPUT_SHEETS,
    SAMPLE_ROW_COUNT,
)
from etabs_mcp.file_io.models import check_cell
from etabs_mcp.file_io.path_validator import FileIOError


def _cell_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "numeric"
    if isinstance(value, (datetime, date, dt_time)):
        return "numeric"
    return "string"


def _detect_header(rows: list[list], has_header: bool | None) -> bool:
    if has_header is not None:
        return has_header
    if len(rows) <= 1:
        return True
    first_row = rows[0]
    data_rows = rows[1:5]
    if not first_row:
        return True
    for col_idx in range(len(first_row)):
        first_type = _cell_type(first_row[col_idx])
        if first_type == "null":
            continue
        type_counts: dict[str, int] = {}
        for row in data_rows:
            if col_idx < len(row):
                t = _cell_type(row[col_idx])
                if t != "null":
                    type_counts[t] = type_counts.get(t, 0) + 1
        if not type_counts:
            continue
        majority_type = max(type_counts, key=type_counts.get)  # type: ignore
        if first_type != majority_type:
            return True
    return False


def _auto_columns(num_cols: int) -> list[str]:
    return [f"col_{i + 1}" for i in range(num_cols)]


class BaseReader(abc.ABC):
    def __init__(self, path: Path) -> None:
        self.path = path
        self._has_header: bool = True
        self._check_file_size()

    def _check_file_size(self) -> None:
        size = self.path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise FileIOError("FILE_TOO_LARGE", f"File is {size:,} bytes; limit is {MAX_FILE_SIZE_BYTES:,} bytes")

    @abc.abstractmethod
    def read(self, *, start_row: int = 0, max_rows: int | None = None, **kwargs: Any) -> Any:
        pass

    @abc.abstractmethod
    def build_summary(self, data: Any) -> dict[str, Any]:
        pass


class CSVReader(BaseReader):
    _CHARDET_MIN_CONFIDENCE = 0.5

    def read(
        self, *, start_row: int = 0, max_rows: int | None = None, has_header: bool | None = None, **kwargs: Any
    ) -> list[list]:
        encoding = self._detect_encoding()
        dialect = self._detect_dialect(encoding)
        all_rows: list[list] = []

        with open(self.path, newline="", encoding=encoding) as f:
            reader = csv.reader(f, dialect)
            try:
                for raw_row in reader:
                    coerced = [_coerce_csv_value(v) for v in raw_row]
                    if len(coerced) > MAX_INPUT_COLUMNS:
                        raise FileIOError("TOO_MANY_COLUMNS", f"Row has {len(coerced)} columns; limit is {MAX_INPUT_COLUMNS}")
                    for cell in coerced:
                        try:
                            check_cell(cell, reject_formula=False)
                        except ValueError as exc:
                            raise FileIOError("INVALID_CELL", f"Invalid cell value: {exc}") from exc
                    all_rows.append(coerced)
            except csv.Error as exc:
                raise FileIOError("CSV_PARSE_ERROR", str(exc)) from None

        self._has_header = _detect_header(all_rows, has_header)

        if self._has_header:
            if not all_rows:
                return []
            data_rows = all_rows[1:]
            if len(data_rows) > MAX_INPUT_ROWS:
                raise FileIOError("TOO_MANY_ROWS", f"File has {len(data_rows)} data rows; limit is {MAX_INPUT_ROWS}")
            sliced = data_rows[start_row:]
            if max_rows is not None:
                sliced = sliced[:max_rows]
            return [all_rows[0], *sliced] if start_row == 0 else sliced
        else:
            if len(all_rows) > MAX_INPUT_ROWS:
                raise FileIOError("TOO_MANY_ROWS", f"File has {len(all_rows)} data rows; limit is {MAX_INPUT_ROWS}")
            sliced = all_rows[start_row:]
            if max_rows is not None:
                sliced = sliced[:max_rows]
            return sliced

    def build_summary(self, data: list[list]) -> dict[str, Any]:
        if self._has_header:
            header = data[0] if data else []
            data_rows = data[1:] if len(data) > 1 else []
        else:
            num_cols = len(data[0]) if data else 0
            header = _auto_columns(num_cols)
            data_rows = data
        return {"total_rows": len(data_rows), "columns": list(header), "sample_rows": [list(r) for r in data_rows[:SAMPLE_ROW_COUNT]]}

    def _detect_encoding(self) -> str:
        raw = self.path.read_bytes()
        result = chardet.detect(raw)
        encoding = result.get("encoding")
        confidence = result.get("confidence", 0)
        if encoding and confidence >= self._CHARDET_MIN_CONFIDENCE:
            return encoding
        return "cp1252"

    def _detect_dialect(self, encoding: str) -> type[csv.Dialect]:
        _COMMON_DELIMITERS = {",", ";", "\t", "|"}
        try:
            with open(self.path, newline="", encoding=encoding) as f:
                sample = f.read(8192)
            dialect = csv.Sniffer().sniff(sample)
            if dialect.delimiter in _COMMON_DELIMITERS:
                return dialect
        except csv.Error:
            pass
        return csv.excel


def _coerce_csv_value(val: str) -> int | float | str:
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        pass
    return val


class XLSXReader(BaseReader):
    def read(self, *, start_row: int = 0, max_rows: int | None = None, sheet: str | None = None, has_header: bool | None = None, **kwargs: Any) -> dict[str, dict[str, Any]]:
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        except Exception as exc:
            raise FileIOError("CORRUPTED_WORKBOOK", f"Cannot open workbook: {exc}") from None
        try:
            self._validate_sheet_count(wb)
            sheets_to_load = self._resolve_sheets(wb, sheet)
            return {name: self._read_sheet(wb[name], name, start_row, max_rows, has_header) for name in sheets_to_load}
        finally:
            wb.close()

    def build_summary(self, data: dict[str, dict[str, Any]]) -> dict[str, Any]:
        sheets = list(data.keys())
        first_sheet = sheets[0] if sheets else None
        first = data[first_sheet] if first_sheet else {"columns": [], "rows": []}
        return {"sheets": sheets, "loaded_sheet": first_sheet, "total_rows": len(first["rows"]), "columns": list(first["columns"]), "sample_rows": [list(r) for r in first["rows"][:SAMPLE_ROW_COUNT]]}

    @staticmethod
    def _validate_sheet_count(wb: Any) -> None:
        if len(wb.sheetnames) > MAX_INPUT_SHEETS:
            raise FileIOError("TOO_MANY_SHEETS", f"Workbook has {len(wb.sheetnames)} sheets; limit is {MAX_INPUT_SHEETS}")

    @staticmethod
    def _resolve_sheets(wb: Any, sheet: str | None) -> list[str]:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise FileIOError("SHEET_NOT_FOUND", f"Sheet '{sheet}' not found in workbook")
            return [sheet]
        return list(wb.sheetnames)

    @staticmethod
    def _read_sheet(ws: Any, name: str, start_row: int, max_rows: int | None, has_header: bool | None) -> dict[str, Any]:
        raw_rows: list[list] = [list(row) for row in ws.iter_rows(values_only=True)]

        if raw_rows and len(raw_rows[0]) > MAX_INPUT_COLUMNS:
            raise FileIOError("TOO_MANY_COLUMNS", f"Sheet '{name}' has {len(raw_rows[0])} columns; limit is {MAX_INPUT_COLUMNS}")

        is_header = _detect_header(raw_rows, has_header)

        all_rows: list[list] = []
        for row in raw_rows:
            converted = [_to_json_primitive(c) for c in row]
            for cell in converted:
                try:
                    check_cell(cell, reject_formula=False)
                except ValueError as exc:
                    raise FileIOError("INVALID_CELL", f"Invalid cell value: {exc}") from exc
            all_rows.append(converted)

        if is_header:
            columns = all_rows[0] if all_rows else []
            data_rows = all_rows[1:]
        else:
            num_cols = len(all_rows[0]) if all_rows else 0
            columns = _auto_columns(num_cols)
            data_rows = all_rows

        if len(data_rows) > MAX_INPUT_ROWS:
            raise FileIOError("TOO_MANY_ROWS", f"Sheet '{name}' exceeds {MAX_INPUT_ROWS} rows")

        sliced = data_rows[start_row:]
        if max_rows is not None:
            sliced = sliced[:max_rows]
        return {"columns": columns, "rows": sliced}


def _to_json_primitive(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.isoformat()
    return str(value)
